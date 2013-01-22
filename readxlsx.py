#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
from openpyxl.reader.excel import load_workbook

wb = load_workbook(filename = sys.argv[1])
sheet_ranges = wb.get_active_sheet()

def valformat(value):
    if type(value) == str:
        pass
    elif type(value) == int:
        value = str(value)
    else:
        pass
    return value

def converthtml(sheet):
    tablestyle = 'border="1" style="border-collapse:collapse"'
    table = 0
    row = 0
    row_empty = 0
    row_max_empty = 20
    end = False
    output = []
    output.append('<table %s>\n'%tablestyle)
    th = None
    while not end:
        row += 1
        col = []
        col_empty = 0
        col_max_empty = 3
        col_end = False
        while not col_end:
            value = sheet.cell(row=row,column=len(col)).value
            if value:
                value = valformat(value)
                col_empty = 0
            else:
                value = ''
                col_empty += 1
            col.append(value)
            if col_empty >= col_max_empty:
                col = col[:-col_max_empty]
                col_end = True
        if len(col) == 1:
            if not th:
                th = col[0]
        elif len(col) > 1:
            if u"开放系统运维组名称未定义 " in col:
                pass
            else:
                if not th:
                    th = sheet.cell(row=row-2,column=0).value
                    output[table] += '  %s<br/>\n'%th
                output[table] += '  <tr>\n'
                output[table] += '    <td style="white-space:nowrap">%s'%('</td>\n    <td style="white-space:nowrap">'.join(col))
                output[table] += '</td>\n  </tr>\n'
                row_empty = 0
        else:
#           output[table] += "  <tr>\n"
#           output[table] += '    <td>%s'%("</td>\n    <td>".join(col))
#           output[table] += '</td>\n  </tr>\n'
            row_empty += 1
        if row_empty >= 2:
            output[table] += '</table>\n'
            if output[table] == '<table %s>\n</table>\n'%tablestyle:
                output.pop()
            else:
                table += 1
            output.append('<table %s>\n'%tablestyle)
            th = None
        if row_empty >= row_max_empty:
            end = True
    output[table] += '</table>\n'
    output.pop()
    return output

tables = converthtml(sheet_ranges)
if sys.argv[2] == 'all':
    for table in tables:
        print table.encode("utf-8")
        print '<br/>'.encode('utf-8')
else:
    for n in sys.argv[2:]:
        print tables[int(n)].encode("utf-8")
        print '<br/>'.encode('utf-8')